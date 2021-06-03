from django.utils import timezone
from django.db import DatabaseError, transaction
from django.core.management.base import BaseCommand
from polls.models import Question
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class Command(BaseCommand):
    help = 'Creates questions and choices from spreadsheet.'

    def add_arguments(self, parser):
        parser.add_argument('path', nargs='+', type=str)

    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename=options['path'][0])
        except InvalidFileException:
            print('Invalid File.')
            return
        except FileNotFoundError:
            print('File not found.')
            return

        ws = wb.active

        row = 2

        try:
            with transaction.atomic():
                while ws.cell(column=1, row=row).value:
                    question = Question(question_text=ws.cell(column=1, row=row).value, pub_date=timezone.now())
                    question.save()
                    col = 2
                    while ws.cell(column=col, row=row).value:
                        question.choice_set.create(choice_text=ws.cell(column=col, row=row).value)
                        col += 1
                    row += 1
        except DatabaseError:
            print('Error creating questions.')

        print('Done creating questions.')
